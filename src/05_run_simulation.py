"""
Script 5: Run Parameter Test Simulation
=========================================
Reads:  db/signals.parquet        (full mode)
        db/signals_quick.parquet  (quick mode)
        db/eq_data.parquet
        config/simulation_params.json

Modes:
  --mode quick  (default for daily run)
        Single fixed param set from quick_run config.
        Output: per-trade PICKS SHEET — one row per trade.
        Sheet name: Pickse
        File  : output/YYYY-MM/QuickRun_Picks_YYYYMMDD_HHMMSS.xlsx
        Extra sheets: MarketData (latest day OHLCV all symbols)
                      BuyHistory (OHLCV for bought stocks, signal→exit date)

  --mode full
        All parameter combinations from param_sweep -> aggregate stats.
        Output: 43-column summary — one row per param combo.
        File  : output/YYYY-MM/Results_YYYYMMDD_HHMMSS.xlsx

SIMULATION RULES:
  - Entry  : D+1 low <= signal_close  -> buy at signal_close price (B0)
  - Additional buys: when buy_count < max_buys (VBA: buyCount < maxBuys) AND next-buy-level >= stop_price
  - Same-day buy and sell NOT allowed (exit checks skip the buy day itself)
  - Invalid (case 1): signal date is the last available date for that symbol (no D+1 data)
  - Invalid (case 2): >10 consecutive calendar days between signal date and next
                      available trading date for that symbol (stock suspended/delisted)
  - Pending: buy not triggered yet, within pending_window_days of signal
  - Expired: buy not triggered, beyond pending_window_days
  - FE-MD  : market_days >= max_duration (trading days counted after first buy)
  - FE-CD  : calendar_days >= force_exit_calendar_days (90 by default)
  - Stop   : based on signal_close (USE_AVGBUY_FOR_STOPLOSS = False)
  - Target : updated on each additional buy (USE_AVGBUY_FOR_TARGET = True)

PICKS SHEET COLUMNS (quick mode — 44 columns for max_buys=2):
  1DChange%, StockName, 5DLow%, 5DLowPrice, RecentLTP,
  BuyDate, BuyClPrice, 5DLowDate, TodayDate,
  BuyCount, AvgBuyPrice, TotalQty, TargetPrice, StoplossPrice, TotalInvestment,
  Order, Status, Duration, DurationGroup, Profits, GainLoss%, Result, ExitType,
  Action, BuyChance, SoldDate, SoldPrice, SoldPrevClose, SoldOpen,
  SoldHigh, SoldLow, SoldClose,
  B0_BoughtDate/PrevClose/Open/High/Low/Close,
  B1_BoughtDate/PrevClose/Open/High/Low/Close

43 AGGREGATE COLUMNS (full mode — exact order):
  Test, DAYSBACK, PCTMIN, PCTMAX, ATHMIN, ATHMAX, MAXBUYS, BUYDROP,
  TARGET, STOPLOSS, MAXDURA, WinRate, TotalTrade, Executed, Open, Closed,
  ProfitTGT, LossSL, LossFEMD, LossFECD, ProfitFEMD, ProfitFECD,
  Pending, Expired, Invalid, TotalRows, Wins, Losses, TotalStock,
  SumProfit, SumGainFin, Dur5, Dur10, Dur15, Dur20, Dur25, Dur30,
  Dur35, Dur40, ExitTGT, ExitSL, ExitFEMD, ExitFECD
"""

import os
import sys
import json
import argparse
import itertools
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG_FILE          = "config/simulation_params.json"
SIGNALS_FILE_FULL    = "db/signals.parquet"
SIGNALS_FILE_QUICK   = "db/signals_quick.parquet"
EQ_FILE              = "db/eq_data.parquet"
SERIES_CHANGED_FILE  = "db/series_changed.parquet"  # symbols that changed away from EQ
OUTPUT_DIR           = "output"

COLUMNS_43 = [
    "Test", "DAYSBACK", "PCTMIN", "PCTMAX", "ATHMIN", "ATHMAX",
    "MAXBUYS", "BUYDROP", "TARGET", "STOPLOSS", "MAXDURA",
    "WinRate", "TotalTrade", "Executed", "Open", "Closed",
    "ProfitTGT", "LossSL", "LossFEMD", "LossFECD", "ProfitFEMD", "ProfitFECD",
    "Pending", "Expired", "Invalid", "TotalRows", "Wins", "Losses", "TotalStock",
    "SumProfit", "SumGainFin",
    "Dur5", "Dur10", "Dur15", "Dur20", "Dur25", "Dur30", "Dur35", "Dur40",
    "ExitTGT", "ExitSL", "ExitFEMD", "ExitFECD"
]

NAVY  = "00203864"
WHITE = "00FFFFFF"
LIGHT = "00F2F2F2"

# Excel number format strings
FMT_DATE  = "DD-MM-YYYY"
FMT_PRICE = "0.00"
FMT_PCT   = "0.00"  # stored as actual percent value (e.g. 10.0), not fraction
FMT_INT   = "General"

# OHLCV sheet columns (Market Data + Buy History sheets)
OHLCV_COLS = ["SYMBOL", "DATE1", "PREV_CLOSE", "OPEN_PRICE",
              "HIGH_PRICE", "LOW_PRICE", "CLOSE_PRICE"]


# ─── CONFIG ───────────────────────────────────────────────────────────────────

def load_config():
    with open(CONFIG_FILE) as f:
        return json.load(f)


def build_trade_combos(cfg, mode):
    if mode == "quick":
        q = cfg["quick_run"]
        return [(q["max_buys"], q["buy_drop"], q["target"],
                 q["stoploss"],  q["max_duration"],
                 q.get("use_stoploss", True), q.get("use_target", True))]
    t = cfg["param_sweep"]["trade"]
    return list(itertools.product(
        t["max_buys"], t["buy_drop"], t["target"],
        t["stoploss"], t["max_duration"],
        t.get("use_stoploss", [True]), t.get("use_target", [True])
    ))


# ─── PRICE DATA ───────────────────────────────────────────────────────────────

def load_series_changed_map():
    """
    Load the series-changed lookup built by 02_filter_eq.py.
    Returns dict: { SYMBOL -> "BE" / "BZ" / "SM" / ... }
    If the file does not exist (02_filter_eq not yet run), returns empty dict.
    """
    if not os.path.exists(SERIES_CHANGED_FILE):
        return {}
    try:
        df = pd.read_parquet(SERIES_CHANGED_FILE)
        return dict(zip(df["SYMBOL"], df["LATEST_SERIES"]))
    except Exception as e:
        print(f"WARN: could not read series_changed.parquet ({e}); skipping series check")
        return {}


def build_price_dict(eq_df):
    print("Building price dictionary...")
    price_dict = {}
    for sym, grp in eq_df.groupby("SYMBOL"):
        grp   = grp.sort_values("DATE1").reset_index(drop=True)
        dates  = grp["DATE1"].values
        closes = grp["CLOSE_PRICE"].values.astype(float)
        highs  = grp["HIGH_PRICE"].values.astype(float)
        lows   = grp["LOW_PRICE"].values.astype(float)
        opens  = grp["OPEN_PRICE"].values.astype(float)
        prevs  = grp["PREV_CLOSE"].values.astype(float)
        day_map = {pd.Timestamp(d): i for i, d in enumerate(dates)}
        price_dict[sym] = {
            "dates":   dates,
            "closes":  closes,
            "highs":   highs,
            "lows":    lows,
            "opens":   opens,
            "prevs":   prevs,
            "day_map": day_map,
        }
    print(f"  Price dict built for {len(price_dict):,} symbols")
    return price_dict


# ─── MARKET DATA HELPERS (for HTML dashboard extra sheets) ───────────────────

def build_latest_market_data(eq_df):
    """
    Extract the latest available OHLCV row per symbol.
    Used for the Market Data tab in the HTML dashboard.
    Returns list of [SYMBOL, DATE1, PREV_CLOSE, OPEN_PRICE, HIGH_PRICE, LOW_PRICE, CLOSE_PRICE]
    """
    print("Extracting latest market data for dashboard...")
    latest_idx = eq_df.groupby("SYMBOL")["DATE1"].idxmax()
    latest_df  = eq_df.loc[latest_idx].sort_values("SYMBOL").reset_index(drop=True)
    rows = []
    for row in latest_df.itertuples(index=False):
        rows.append([
            str(row.SYMBOL),
            row.DATE1.to_pydatetime().replace(tzinfo=None),
            round(float(row.PREV_CLOSE),  2),
            round(float(row.OPEN_PRICE),  2),
            round(float(row.HIGH_PRICE),  2),
            round(float(row.LOW_PRICE),   2),
            round(float(row.CLOSE_PRICE), 2),
        ])
    print(f"  Latest market data: {len(rows):,} symbols")
    return rows


def build_buy_history(eq_df, bought_ranges):
    """
    Extract OHLCV for bought stocks from (signal_date - 7 days) to exit/last date.
    bought_ranges: list of (symbol_str, start_pd_Timestamp, end_pd_Timestamp)
    Returns list of [SYMBOL, DATE1, PREV_CLOSE, OPEN_PRICE, HIGH_PRICE, LOW_PRICE, CLOSE_PRICE]
    Used for the Stock History tab in the HTML dashboard.
    Performance: pre-groups eq_df by SYMBOL to avoid O(N×M) full-scan per trade.
    """
    if not bought_ranges:
        return []
    print(f"Extracting buy history for {len(bought_ranges):,} bought stocks...")
    # Pre-group by SYMBOL once — O(N) instead of O(N×M) per trade
    sym_groups = {sym: grp for sym, grp in eq_df.groupby("SYMBOL")}
    rows = []
    for sym, start_ts, end_ts in bought_ranges:
        if sym not in sym_groups:
            continue
        grp = sym_groups[sym]
        sub = grp[(grp["DATE1"] >= start_ts) & (grp["DATE1"] <= end_ts)]
        if sub.empty:
            continue
        for row in sub.sort_values("DATE1").itertuples(index=False):
            rows.append([
                sym,
                row.DATE1.to_pydatetime().replace(tzinfo=None),
                round(float(row.PREV_CLOSE),  2),
                round(float(row.OPEN_PRICE),  2),
                round(float(row.HIGH_PRICE),  2),
                round(float(row.LOW_PRICE),   2),
                round(float(row.CLOSE_PRICE), 2),
            ])
    print(f"  Buy history rows: {len(rows):,}")
    return rows


# ─── DURATION GROUP ───────────────────────────────────────────────────────────

def duration_group(market_days):
    """Return DurationGroup: integer bucket of 5 up to 50, '50+' for >50."""
    if market_days is None:
        return None
    if market_days > 50:
        return "50+"
    return ((market_days - 1) // 5 + 1) * 5


# ─── SIMULATION (DETAILED — Picks Sheet / Quick Run) ─────────────────────────

def simulate_trade_detailed(sym, signal_date, signal_close, price_dict,
                             max_buys, buy_drop, target_pct, stoploss_pct,
                             max_duration, investment_per_buy,
                             force_exit_calendar_days, pending_window_days,
                             global_last_date=None,
                             use_stoploss=True, use_target=True):
    """
    Simulate one trade and return all per-trade details for the picks sheet.

    Key rules:
      - If start_idx >= last_idx  (signal date is the last available date for
        this symbol) -> INVALID (no D+1 data to attempt entry).
      - If gap between signal date and next available date > 10 calendar days
        -> INVALID (stock suspended / delisted / data missing 10+ days).
      - Loop: range(start_idx+1, last_idx)  -- excludes the very last data day
        ("today") from both buy attempts and exit checks.
      - Same-day buy and sell NOT allowed: exit checks skip the buy day itself.
      - max_buys = TOTAL buy slots (VBA semantics: buyCount < max_buys);
        total buy blocks = max_buys (B0..B(max_buys-1))
    """
    invalid = {
        "order": "Invalid", "status": None, "action": "Skip",
        "buy_count": None, "avg_buy_price": None, "total_qty": None,
        "total_investment": None, "target_price": None, "stop_price": None,
        "first_buy_date": None, "exit_found": False,
        "exit_date": None, "exit_price": None, "exit_type": None,
        "profit": None, "gain_pct": None, "market_days": None,
        "result_str": "Invalid: No market data found",
        "duration_group": None,
        "buy_chance": None,
        "sold_prev_close": None, "sold_open": None,
        "sold_high": None, "sold_low": None, "sold_close": None,
        "buys": [],
        "had_buy_chance": False,
    }

    if sym not in price_dict:
        return invalid

    # ── INVALID case 0: symbol's series changed away from EQ ─────────────────
    # e.g. FAZE3Q was EQ but latest bhav shows SERIES=BE → not tradeable as EQ
    series_changed_map = price_dict.get("__series_changed__", {})
    if sym in series_changed_map:
        inv0 = dict(invalid)
        inv0["result_str"] = f"Invalid: Series changed from EQ to {series_changed_map[sym]}"
        return inv0

    pd_data  = price_dict[sym]
    dates    = pd_data["dates"]
    closes   = pd_data["closes"]
    highs    = pd_data["highs"]
    lows     = pd_data["lows"]
    opens    = pd_data["opens"]
    prevs    = pd_data["prevs"]
    day_map  = pd_data["day_map"]

    sig_ts = pd.Timestamp(signal_date)
    if sig_ts not in day_map:
        return invalid

    start_idx = day_map[sig_ts]
    last_idx  = len(dates) - 1

    # ── INVALID case 1: signal on last available date — no D+1 data ──────────
    # VBA: stockDict does not have signal date key → "Invalid: No data on signal date"
    if start_idx >= last_idx:
        inv1 = dict(invalid)
        inv1["result_str"] = "Invalid: No data on signal date"
        return inv1

    # ── INVALID case 2: >10 consecutive calendar days gap to next data ────────
    # VBA: hasDataGap = True → "Invalid: Data gap detected"
    next_avail = pd.Timestamp(dates[start_idx + 1])
    if (next_avail - sig_ts).days > 10:
        inv2 = dict(invalid)
        inv2["result_str"] = "Invalid: Data gap detected"
        return inv2

    # ── INVALID case 3: last data is stale (stock stopped trading) ─────────────
    # VBA: gapFromLastData > MAX_MISSING_DAYS → "Invalid: Stale data"
    last_avail_ts = pd.Timestamp(dates[last_idx])
    today_ts      = pd.Timestamp("today").normalize()
    if (today_ts - last_avail_ts).days > 10:
        inv3 = dict(invalid)
        inv3["result_str"] = "Invalid: Stale data"
        return inv3

    # Prices for target and stop (based on signal_close per VBA logic)
    stop_price   = round(signal_close * (1 - stoploss_pct), 2)
    target_price = round(signal_close * (1 + target_pct),  2)

    buy_count        = 0
    avg_buy_price    = 0.0
    total_qty        = 0
    total_investment = 0.0
    first_buy_date   = None
    buy_day_indices  = set()
    market_days      = 0
    exit_found       = False
    exit_type        = None
    exit_price       = 0.0
    exit_date        = None
    exit_idx         = -1
    buys             = []
    had_buy_chance   = False

    # ── Main loop — D+1 through last available date (inclusive) ─────────────
    # BUG FIX: was range(start_idx+1, last_idx) — excluded the final data point.
    # That caused buy/exit triggers on the last day to be missed (e.g. RecentLTP
    # already below BuyClPrice yet Status=Pending).  Now includes every day.
    prev_date_ts = pd.Timestamp(dates[start_idx])  # track previous data date for mid-period gap check
    for i in range(start_idx + 1, len(dates)):
        curr_ts  = pd.Timestamp(dates[i])

        # ── BUG FIX: Mid-period data gap detection (matches VBA logic) ───────
        # VBA scans consecutiveMissingDays throughout entire holding period.
        # If gap between any two consecutive data points > 10 calendar days
        # after first buy, mark Invalid: Data gap detected.
        if buy_count > 0 and (curr_ts - prev_date_ts).days > 10:
            return {
                **invalid,
                "order":      "Invalid",
                "result_str": "Invalid: Data gap detected",
            }
        prev_date_ts = curr_ts

        low_px   = float(lows[i])
        high_px  = float(highs[i])
        close_px = float(closes[i])
        open_px  = float(opens[i])
        prev_px  = float(prevs[i])

        if buy_count == 0:
            # ── First buy (B0): day low touches signal_close ─────────────────
            if low_px <= signal_close:
                buy_count      = 1
                first_buy_date = curr_ts
                buy_day_indices.add(i)
                qty = int(investment_per_buy / signal_close)
                if qty < 1:
                    qty = 1
                total_qty        = qty
                total_investment = signal_close * qty
                avg_buy_price    = round(total_investment / total_qty, 2)
                # Target updates on each buy (USE_AVGBUY_FOR_TARGET = True)
                target_price     = round(avg_buy_price * (1 + target_pct), 2)
                buys.append({
                    "date":       curr_ts,
                    "prev_close": prev_px,
                    "open":       open_px,
                    "high":       high_px,
                    "low":        low_px,
                    "close":      close_px,
                })
        else:
            is_buy_day = i in buy_day_indices
            if not is_buy_day:
                market_days += 1

            cal_days = (curr_ts - first_buy_date).days if first_buy_date else 0

            # ── Check exits (only on non-buy days — same-day sell not allowed) ─
            if not is_buy_day:
                if use_stoploss and low_px <= stop_price:
                    exit_found = True
                    exit_price = stop_price
                    exit_type  = "Stoploss Triggered"
                    exit_date  = curr_ts
                    exit_idx   = i
                    break
                if use_target and high_px >= target_price:
                    exit_found = True
                    exit_price = target_price
                    exit_type  = "Target Achieved"
                    exit_date  = curr_ts
                    exit_idx   = i
                    break
                if market_days >= max_duration:
                    exit_found = True
                    exit_price = round(close_px, 2)
                    exit_type  = "Force Exit - Market Days"
                    exit_date  = curr_ts
                    exit_idx   = i
                    break
                if cal_days >= force_exit_calendar_days:
                    exit_found = True
                    exit_price = round(close_px, 2)
                    exit_type  = "Force Exit - Calendar Days"
                    exit_date  = curr_ts
                    exit_idx   = i
                    break

            # ── Additional buys (VBA: buyCount < maxBuys) ─────────────────────
            if not exit_found and not is_buy_day and buy_count < max_buys:
                buy_level = round(avg_buy_price * (1 - buy_drop), 2)
                if low_px <= buy_level:
                    had_buy_chance = True
                    if buy_level >= stop_price:
                        buy_count        += 1
                        buy_day_indices.add(i)
                        qty = int(investment_per_buy / buy_level)
                        if qty < 1:
                            qty = 1
                        total_qty        += qty
                        total_investment += buy_level * qty
                        avg_buy_price     = round(total_investment / total_qty, 2)
                        target_price      = round(avg_buy_price * (1 + target_pct), 2)
                        buys.append({
                            "date":       curr_ts,
                            "prev_close": prev_px,
                            "open":       open_px,
                            "high":       high_px,
                            "low":        low_px,
                            "close":      close_px,
                        })

    # ── Determine order / status / action ─────────────────────────────────────
    if buy_count == 0:
        # BUG FIX: was using symbol's own last date (dates[last_idx]).
        # If a stock was delisted/suspended after just a few days, its last date
        # was only days after signal → wrongly showed as Pending for months.
        # Fix: always use the GLOBAL last data date for the Pending/Expired decision.
        ref_ts    = pd.Timestamp(global_last_date) if global_last_date else pd.Timestamp(dates[last_idx])
        days_from = (ref_ts - sig_ts).days
        if days_from <= pending_window_days:
            order = "Pending"
            action = "Buy"
        else:
            order = "Expired"
            action = "Skip"
        return {
            "order": order,
            "status": "Not Triggered" if order == "Pending" else "Expired",
            "action": action,
            "buy_count": 0, "avg_buy_price": None, "total_qty": None,
            "total_investment": None,
            "target_price": target_price, "stop_price": stop_price,
            "first_buy_date": None, "exit_found": False,
            "exit_date": None, "exit_price": None, "exit_type": None,
            "profit": None, "gain_pct": None, "market_days": None,
            "result_str": None, "duration_group": None,
            "buy_chance": None,
            "sold_prev_close": None, "sold_open": None,
            "sold_high": None, "sold_low": None, "sold_close": None,
            "buys": [], "had_buy_chance": False,
        }

    # ── Executed — Open ───────────────────────────────────────────────────────
    if not exit_found:
        return {
            "order": "Executed", "status": "Open", "action": "Hold",
            "buy_count": buy_count, "avg_buy_price": avg_buy_price,
            "total_qty": total_qty, "total_investment": round(total_investment, 2),
            "target_price": target_price, "stop_price": stop_price,
            "first_buy_date": first_buy_date, "exit_found": False,
            "exit_date": None, "exit_price": None, "exit_type": None,
            "profit": None,   # computed in build_picks_row using recent_ltp
            "gain_pct": None, # computed in build_picks_row using recent_ltp
            "market_days": market_days,
            "result_str": None,
            "duration_group": duration_group(market_days),
            "buy_chance": "Buy Chance" if had_buy_chance else None,
            "sold_prev_close": None, "sold_open": None,
            "sold_high": None, "sold_low": None, "sold_close": None,
            "buys": buys, "had_buy_chance": had_buy_chance,
        }

    # ── Executed — Closed ─────────────────────────────────────────────────────
    profit   = round((exit_price - avg_buy_price) * total_qty, 2)
    gain_pct = round(((exit_price - avg_buy_price) / avg_buy_price) * 100, 2) \
               if avg_buy_price > 0 else 0.0

    if   exit_type == "Target Achieved":           result_str = "Profit-TGT"
    elif exit_type == "Stoploss Triggered":         result_str = "Loss-SL"
    elif exit_type == "Force Exit - Market Days":   result_str = ("Profit" if profit >= 0 else "Loss") + "-FE-MD"
    elif exit_type == "Force Exit - Calendar Days": result_str = ("Profit" if profit >= 0 else "Loss") + "-FE-CD"
    else:                                           result_str = "Profit" if profit >= 0 else "Loss"

    sold_prev_close = sold_open = sold_high = sold_low = sold_close = None
    if exit_idx >= 0:
        sold_prev_close = round(float(prevs[exit_idx]),  2)
        sold_open       = round(float(opens[exit_idx]),  2)
        sold_high       = round(float(highs[exit_idx]),  2)
        sold_low        = round(float(lows[exit_idx]),   2)
        sold_close      = round(float(closes[exit_idx]), 2)

    return {
        "order": "Executed", "status": "Closed", "action": "Exit",
        "buy_count": buy_count, "avg_buy_price": avg_buy_price,
        "total_qty": total_qty, "total_investment": round(total_investment, 2),
        "target_price": target_price, "stop_price": stop_price,
        "first_buy_date": first_buy_date, "exit_found": True,
        "exit_date": exit_date, "exit_price": exit_price, "exit_type": exit_type,
        "profit": profit, "gain_pct": gain_pct,
        "market_days": market_days,
        "result_str": result_str,
        "duration_group": duration_group(market_days),
        "buy_chance": "Buy Chance" if had_buy_chance else None,
        "sold_prev_close": sold_prev_close, "sold_open": sold_open,
        "sold_high": sold_high, "sold_low": sold_low, "sold_close": sold_close,
        "buys": buys, "had_buy_chance": had_buy_chance,
    }


# ─── SIMULATION (SIMPLE — Full mode aggregate stats) ─────────────────────────

def simulate_trade(sym, signal_date, signal_close, price_dict,
                   max_buys, buy_drop, target_pct, stoploss_pct,
                   max_duration, investment_per_buy,
                   force_exit_calendar_days, pending_window_days,
                   global_last_date=None,
                   use_stoploss=True, use_target=True):
    """Lightweight simulation for full-mode parameter sweep aggregate stats."""
    if sym not in price_dict:
        return {"order": "Invalid"}

    # ── INVALID case 0: symbol's series changed away from EQ ─────────────────
    series_changed_map = price_dict.get("__series_changed__", {})
    if sym in series_changed_map:
        return {"order": "Invalid", "result_str": f"Invalid: Series changed from EQ to {series_changed_map[sym]}"}

    pd_data  = price_dict[sym]
    dates    = pd_data["dates"]
    closes   = pd_data["closes"]
    highs    = pd_data["highs"]
    lows     = pd_data["lows"]
    day_map  = pd_data["day_map"]

    sig_ts = pd.Timestamp(signal_date)
    if sig_ts not in day_map:
        return {"order": "Invalid"}

    start_idx = day_map[sig_ts]
    last_idx  = len(dates) - 1

    # ── INVALID case 1: signal on last available date ─────────────────────────
    if start_idx >= last_idx:
        return {"order": "Invalid", "result_str": "Invalid: No data on signal date"}

    # ── INVALID case 2: >10 consecutive calendar days gap to next data ────────
    next_avail = pd.Timestamp(dates[start_idx + 1])
    if (next_avail - sig_ts).days > 10:
        return {"order": "Invalid", "result_str": "Invalid: Data gap detected"}

    # ── INVALID case 3: stale data (stock stopped trading) ───────────────────
    last_avail_ts2 = pd.Timestamp(dates[last_idx])
    today_ts2      = pd.Timestamp("today").normalize()
    if (today_ts2 - last_avail_ts2).days > 10:
        return {"order": "Invalid", "result_str": "Invalid: Stale data"}

    stop_price   = round(signal_close * (1 - stoploss_pct), 2)
    target_price = round(signal_close * (1 + target_pct),  2)

    buy_count        = 0
    avg_buy_price    = 0.0
    total_qty        = 0
    total_investment = 0.0
    first_buy_date   = None
    buy_day_indices  = set()
    market_days      = 0
    exit_found       = False
    exit_type        = None
    exit_price       = 0.0
    exit_date        = None

    prev_date_ts2 = pd.Timestamp(dates[start_idx])  # track previous data date for mid-period gap check
    for i in range(start_idx + 1, len(dates)):   # BUG FIX: include last data point
        curr_ts  = pd.Timestamp(dates[i])

        # ── BUG FIX: Mid-period data gap detection (matches VBA logic) ───────
        if buy_count > 0 and (curr_ts - prev_date_ts2).days > 10:
            return {"order": "Invalid", "result_str": "Invalid: Data gap detected"}
        prev_date_ts2 = curr_ts
        low_px   = float(lows[i])
        high_px  = float(highs[i])
        close_px = float(closes[i])

        if buy_count == 0:
            if low_px <= signal_close:
                buy_count      = 1
                first_buy_date = curr_ts
                buy_day_indices.add(i)
                qty = int(investment_per_buy / signal_close)
                if qty < 1: qty = 1
                total_qty        = qty
                total_investment = signal_close * qty
                avg_buy_price    = round(total_investment / total_qty, 2)
                target_price     = round(avg_buy_price * (1 + target_pct), 2)
        else:
            is_buy_day = i in buy_day_indices
            if not is_buy_day:
                market_days += 1
            cal_days = (curr_ts - first_buy_date).days if first_buy_date else 0
            if not is_buy_day:
                if use_stoploss and low_px <= stop_price:
                    exit_found = True; exit_price = stop_price
                    exit_type  = "Stoploss Triggered"; exit_date = curr_ts; break
                if use_target and high_px >= target_price:
                    exit_found = True; exit_price = target_price
                    exit_type  = "Target Achieved"; exit_date = curr_ts; break
                if market_days >= max_duration:
                    exit_found = True; exit_price = round(close_px, 2)
                    exit_type  = "Force Exit - Market Days"; exit_date = curr_ts; break
                if cal_days >= force_exit_calendar_days:
                    exit_found = True; exit_price = round(close_px, 2)
                    exit_type  = "Force Exit - Calendar Days"; exit_date = curr_ts; break
            if not exit_found and not is_buy_day and buy_count < max_buys:
                buy_level = round(avg_buy_price * (1 - buy_drop), 2)
                if buy_level >= stop_price and low_px <= buy_level:
                    buy_count        += 1
                    buy_day_indices.add(i)
                    qty = int(investment_per_buy / buy_level)
                    if qty < 1: qty = 1
                    total_qty        += qty
                    total_investment += buy_level * qty
                    avg_buy_price     = round(total_investment / total_qty, 2)
                    target_price      = round(avg_buy_price * (1 + target_pct), 2)

    if buy_count == 0:
        # BUG FIX: use global last date (not symbol's last date) for Pending/Expired
        ref_ts    = pd.Timestamp(global_last_date) if global_last_date else pd.Timestamp(dates[last_idx])
        days_from = (ref_ts - sig_ts).days
        order = "Pending" if days_from <= pending_window_days else "Expired"
        return {"order": order, "status": order, "profit": 0, "gain_pct": 0,
                "exit_type": None, "result": None, "market_days": 0}

    if not exit_found:
        return {"order": "Executed", "status": "Open", "profit": 0, "gain_pct": 0,
                "exit_type": None, "result": None, "market_days": market_days}

    profit   = round((exit_price - avg_buy_price) * total_qty, 2)
    gain_pct = round(((exit_price - avg_buy_price) / avg_buy_price) * 100, 2) \
               if avg_buy_price > 0 else 0

    if   exit_type == "Target Achieved":           result = "Profit-TGT"
    elif exit_type == "Stoploss Triggered":         result = "Loss-SL"
    elif exit_type == "Force Exit - Market Days":   result = ("Profit" if profit >= 0 else "Loss") + "-FE-MD"
    elif exit_type == "Force Exit - Calendar Days": result = ("Profit" if profit >= 0 else "Loss") + "-FE-CD"
    else:                                           result = "Profit" if profit >= 0 else "Loss"

    return {
        "order": "Executed", "status": "Closed",
        "profit": profit, "gain_pct": gain_pct,
        "exit_type": exit_type, "result": result,
        "market_days": market_days,
    }


# ─── AGGREGATE STATS (full mode) ─────────────────────────────────────────────

def aggregate_stats(results):
    wins=0; losses=0
    c_profit_tgt=0; c_loss_sl=0
    c_loss_femd=0;  c_profit_femd=0
    c_loss_fecd=0;  c_profit_fecd=0
    dur5=0; dur10=0; dur15=0; dur20=0
    dur25=0; dur30=0; dur35=0; dur40=0
    pending=0; expired=0; invalid=0
    exec_count=0; open_count=0; closed_count=0
    sum_profit=0.0; sum_gain=0.0
    exit_tgt=0; exit_sl=0; exit_femd=0; exit_fecd=0

    for r in results:
        order = r.get("order", "Invalid")
        if   order == "Pending":  pending += 1
        elif order == "Expired":  expired += 1
        elif order == "Invalid":  invalid += 1
        elif order == "Executed":
            exec_count += 1
            status = r.get("status", "")
            if   status == "Open":   open_count += 1
            elif status == "Closed":
                closed_count += 1
                result    = r.get("result",     "") or ""
                exit_type = r.get("exit_type",  "") or ""
                profit    = r.get("profit",     0)  or 0
                gain_pct  = r.get("gain_pct",   0)  or 0
                mdays     = r.get("market_days",0)  or 0

                sum_profit += profit
                sum_gain   += gain_pct

                ru = result.upper()
                if   "PROFIT-TGT"  in ru: c_profit_tgt  += 1; wins   += 1
                elif "LOSS-SL"     in ru: c_loss_sl      += 1; losses += 1
                elif ru == "PROFIT-FE-MD": c_profit_femd += 1; wins   += 1
                elif ru == "LOSS-FE-MD":   c_loss_femd   += 1; losses += 1
                elif ru == "PROFIT-FE-CD": c_profit_fecd += 1; wins   += 1
                elif ru == "LOSS-FE-CD":   c_loss_fecd   += 1; losses += 1
                elif "PROFIT" in ru:       wins   += 1
                else:                      losses += 1

                eu = exit_type.upper()
                if   "TARGET ACHIEVED"     in eu: exit_tgt  += 1
                elif "STOPLOSS TRIGGERED"  in eu: exit_sl   += 1
                elif "FORCE EXIT - MARKET" in eu: exit_femd += 1
                elif "FORCE EXIT - CALEND" in eu: exit_fecd += 1

                dg = duration_group(mdays)
                dg_num = dg if isinstance(dg, int) else 40
                if   dg_num <= 5:  dur5  += 1
                elif dg_num <= 10: dur10 += 1
                elif dg_num <= 15: dur15 += 1
                elif dg_num <= 20: dur20 += 1
                elif dg_num <= 25: dur25 += 1
                elif dg_num <= 30: dur30 += 1
                elif dg_num <= 35: dur35 += 1
                else:              dur40 += 1

    total_stocks = len(results)
    win_rate     = round((wins / closed_count * 100) if closed_count > 0 else 0, 2)

    return {
        "WinRate":    win_rate,
        "TotalTrade": exec_count + pending + expired,
        "Executed":   exec_count,
        "Open":       open_count,
        "Closed":     closed_count,
        "ProfitTGT":  c_profit_tgt, "LossSL":     c_loss_sl,
        "LossFEMD":   c_loss_femd,  "LossFECD":   c_loss_fecd,
        "ProfitFEMD": c_profit_femd,"ProfitFECD": c_profit_fecd,
        "Pending":    pending,       "Expired":    expired,
        "Invalid":    invalid,       "TotalRows":  total_stocks,
        "Wins":       wins,          "Losses":     losses,
        "TotalStock": total_stocks,  "SumProfit":  round(sum_profit, 2),
        "SumGainFin": round(sum_gain, 4),
        "Dur5":  dur5,  "Dur10": dur10, "Dur15": dur15, "Dur20": dur20,
        "Dur25": dur25, "Dur30": dur30, "Dur35": dur35, "Dur40": dur40,
        "ExitTGT":  exit_tgt,  "ExitSL":   exit_sl,
        "ExitFEMD": exit_femd, "ExitFECD": exit_fecd,
    }


# ─── PICKS SHEET HELPERS ─────────────────────────────────────────────────────

def get_picks_columns(max_buys):
    """
    Column headers for Picks Sheet.
    VBA semantics: total buy blocks = max_buys (B0..B(max_buys-1))
    """
    cols = [
        "1DChange%", "StockName", "5DLow%", "5DLowPrice", "RecentLTP",
        "BuyDate", "BuyClPrice", "5DLowDate", "TodayDate",
        "BuyCount", "AvgBuyPrice", "TotalQty", "TargetPrice", "StoplossPrice",
        "TotalInvestment", "Order", "Status", "Duration", "DurationGroup",
        "Profits", "GainLoss%", "Result", "ExitType", "Action", "BuyChance",
        "SoldDate", "SoldPrice", "SoldPrevClose", "SoldOpen",
        "SoldHigh", "SoldLow", "SoldClose",
    ]
    for b in range(max_buys):       # B0 through B(max_buys-1)
        cols += [
            f"B{b}_BoughtDate", f"B{b}_PrevClose", f"B{b}_Open",
            f"B{b}_High",       f"B{b}_Low",        f"B{b}_Close",
        ]
    return cols


def _col_fmt(col_name):
    """Return openpyxl number format string for a column name."""
    if "Date" in col_name or col_name.endswith("Date"):
        return FMT_DATE
    if col_name in ("1DChange%", "5DLow%", "GainLoss%"):
        return FMT_PCT
    if col_name in ("BuyCount", "TotalQty", "Duration", "DurationGroup", "BuyChance",
                    "Order", "Status", "Result", "ExitType", "Action",
                    "StockName", "WinRate"):
        return FMT_INT
    # Price / money columns
    if col_name in ("5DLowPrice", "RecentLTP", "BuyClPrice", "AvgBuyPrice",
                    "TargetPrice", "StoplossPrice", "TotalInvestment",
                    "Profits", "SoldPrice", "SoldPrevClose", "SoldOpen",
                    "SoldHigh", "SoldLow", "SoldClose"):
        return FMT_PRICE
    # Buy-block price columns  (B0_PrevClose, B0_Open, etc.)
    if "_PrevClose" in col_name or "_Open" in col_name or "_High" in col_name \
       or "_Low" in col_name or "_Close" in col_name:
        return FMT_PRICE
    return FMT_INT


def _to_dt(ts):
    """Convert to Python datetime or None."""
    if ts is None:
        return None
    try:
        t = pd.Timestamp(ts)
        if pd.isna(t):
            return None
        return t.to_pydatetime().replace(tzinfo=None)
    except Exception:
        return None


def build_picks_row(sig, sim, price_dict, max_buys, last_data_date):
    """
    Assemble one picks-sheet row from a signal dict and simulation result.
    All date fields are Python datetime objects (formatted DD-MM-YYYY by writer).
    Price fields are rounded to 2 decimal places.
    GainLoss%, 1DChange%, 5DLow% are stored as actual percentages (e.g. -8.0),
    formatted as percentage in Excel via 0.00% format.
    """
    sym   = str(sig["SYMBOL"])
    order = sim["order"]

    # RecentLTP = latest available close for this symbol
    recent_ltp = None
    if sym in price_dict:
        recent_ltp = round(float(price_dict[sym]["closes"][-1]), 2)
    if recent_ltp is None:
        recent_ltp = round(float(sig["SIGNAL_CLOSE"]), 2)

    # Date fields as Python datetime objects
    buy_date   = _to_dt(sig["SIGNAL_DATE"])
    min5d_date = _to_dt(sig["MIN_5D_DATE"])
    today_date = last_data_date if order != "Invalid" else None
    sold_date  = _to_dt(sim["exit_date"]) if sim["exit_found"] else None

    # BuyClPrice — the signal-day close price
    buy_cl_price = round(float(sig["SIGNAL_CLOSE"]), 2)

    # BuyChance — VBA: "Buy Chance" if recent LTP <= signal close price
    # (= current price is still at or below original entry signal → still a buying opportunity)
    # VBA: `If CDbl(ltp) <= signalPrice Then buyChanceValue = "Buy Chance"`
    # Applied to all non-Invalid rows (Pending, Expired, Executed-Open, Executed-Closed)
    if order != "Invalid" and recent_ltp is not None and recent_ltp <= buy_cl_price:
        buy_chance_val = "Buy Chance"
    else:
        buy_chance_val = None

    # Unrealized P&L for Open trades (computed using recent_ltp)
    profit   = sim["profit"]
    gain_pct = sim["gain_pct"]
    if order == "Executed" and sim["status"] == "Open":
        avg_buy = sim["avg_buy_price"]
        qty     = sim["total_qty"]
        if avg_buy and avg_buy > 0 and recent_ltp and qty:
            profit   = round((recent_ltp - avg_buy) * qty, 2)
            gain_pct = round(((recent_ltp - avg_buy) / avg_buy) * 100, 2)

    row = [
        round(float(sig["PCT_1D_CHANGE"]) * 100, 2),  # 1DChange% (actual percent, matches VBA)
        sym,                                    # StockName
        round(float(sig["PCT_FROM_LOW"]) * 100, 2),   # 5DLow%     (actual percent, matches VBA)
        round(float(sig.get("MIN_5D_LOW") or sig.get("MIN_5D_CLOSE") or 0), 2),  # 5DLowPrice
        recent_ltp,                             # RecentLTP
        buy_date,                               # BuyDate (datetime)
        buy_cl_price,                           # BuyClPrice
        min5d_date,                             # 5DLowDate (datetime)
        today_date,                             # TodayDate (datetime or None)
        sim["buy_count"],                       # BuyCount
        sim["avg_buy_price"],                   # AvgBuyPrice
        sim["total_qty"],                       # TotalQty
        sim["target_price"],                    # TargetPrice
        sim["stop_price"],                      # StoplossPrice
        sim["total_investment"],                # TotalInvestment
        sim["order"],                           # Order
        sim["status"],                          # Status
        sim["market_days"],                     # Duration
        sim["duration_group"],                  # DurationGroup
        profit,                                 # Profits
        gain_pct,                               # GainLoss% (decimal fraction)
        sim["result_str"],                      # Result
        sim["exit_type"],                       # ExitType
        sim["action"],                          # Action
        buy_chance_val,                         # BuyChance (VBA: recent_ltp <= signal_close)
        sold_date,                              # SoldDate (datetime)
        sim["exit_price"],                      # SoldPrice
        sim["sold_prev_close"],                 # SoldPrevClose
        sim["sold_open"],                       # SoldOpen
        sim["sold_high"],                       # SoldHigh
        sim["sold_low"],                        # SoldLow
        sim["sold_close"],                      # SoldClose
    ]

    # Per-buy blocks B0..B(max_buys-1) — VBA: buyCount < max_buys
    buys = sim.get("buys", [])
    for b in range(max_buys):
        if b < len(buys):
            bdata = buys[b]
            row += [
                _to_dt(bdata["date"]),
                round(float(bdata["prev_close"]), 2),
                round(float(bdata["open"]),       2),
                round(float(bdata["high"]),       2),
                round(float(bdata["low"]),        2),
                round(float(bdata["close"]),      2),
            ]
        else:
            row += [None, None, None, None, None, None]

    return row


def _style_ohlcv_header(ws):
    """Apply navy header formatting to an OHLCV sheet."""
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col_name in enumerate(OHLCV_COLS, 1):
        cell           = ws.cell(row=1, column=ci)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        # Date column
        if col_name == "DATE1":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=ci).number_format = FMT_DATE
        # Price columns
        elif col_name not in ("SYMBOL",):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=ci).number_format = FMT_PRICE
    ws.column_dimensions["A"].width = 16  # SYMBOL
    ws.column_dimensions["B"].width = 14  # DATE1
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 12
    ws.freeze_panes = ws["A2"]


def write_picks_excel(rows, columns, out_path,
                      market_data=None, buy_history=None):
    """
    Write picks sheet to Excel.
      Sheet name : Pickse
      Row 1      : Bold header — navy background, white text
      Row 2+     : Data with alternating row fill
      Formats    : prices = 0.00 | percentages = 0.00% | dates = DD-MM-YYYY
      Extra sheets (if provided):
        MarketData  — latest day OHLCV for all EQ symbols (Market Data tab)
        BuyHistory  — OHLCV from signal→exit for bought stocks (Stock History tab)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pickse"

    thin    = Side(style="thin", color="BFBFBF")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols   = len(columns)

    # Pre-compute per-column format strings
    col_fmts = [_col_fmt(c) for c in columns]

    # ── Header row (row 1) ────────────────────────────────────────────────────
    ws.append(columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws.row_dimensions[1].height = 18

    # ── Data rows (row 2+) ────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows):
        ws.append(row)
        fill_color  = LIGHT if (r_idx % 2 == 0) else "00FFFFFF"
        actual_row  = r_idx + 2
        for col_idx in range(1, ncols + 1):
            cell               = ws.cell(row=actual_row, column=col_idx)
            cell.fill          = PatternFill("solid", fgColor=fill_color)
            cell.border        = bdr
            cell.alignment     = Alignment(horizontal="center")
            cell.number_format = col_fmts[col_idx - 1]

    # ── Freeze panes & auto-filter ────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = (ws.cell(row=1, column=1).coordinate + ":" +
                          ws.cell(row=1, column=ncols).coordinate)

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, 1):
        w = 13
        if col_name == "StockName":                    w = 16
        elif "Date" in col_name:                       w = 13
        elif col_name in ("Result", "ExitType"):       w = 22
        elif col_name in ("Order", "Status", "Action"): w = 14
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Page setup ────────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    # ── Extra sheet: MarketData ───────────────────────────────────────────────
    if market_data:
        ws_md = wb.create_sheet("MarketData")
        ws_md.append(OHLCV_COLS)
        for r in market_data:
            ws_md.append(r)
        _style_ohlcv_header(ws_md)

    # ── Extra sheet: BuyHistory ───────────────────────────────────────────────
    if buy_history:
        ws_bh = wb.create_sheet("BuyHistory")
        ws_bh.append(OHLCV_COLS)
        for r in buy_history:
            ws_bh.append(r)
        _style_ohlcv_header(ws_bh)

    wb.save(out_path)


# ─── STYLE (full mode aggregate sheet) ───────────────────────────────────────

def style_sheet(ws, mode_label):
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(COLUMNS_43))
    tc          = ws.cell(row=1, column=1)
    tc.value    = (f"NSE Simulation Results [{mode_label}] — "
                   f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}")
    tc.font     = Font(bold=True, size=14, color=WHITE)
    tc.fill     = PatternFill("solid", fgColor=NAVY)
    tc.alignment= Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, col_name in enumerate(COLUMNS_43, 1):
        cell           = ws.cell(row=2, column=col_idx)
        cell.value     = col_name
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws.row_dimensions[2].height = 18

    last_row = ws.max_row
    for row in range(3, last_row + 1):
        fill_color = LIGHT if row % 2 == 0 else "00FFFFFF"
        for col in range(1, len(COLUMNS_43) + 1):
            cell           = ws.cell(row=row, column=col)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=3, column=1)
    ws.auto_filter.ref = (ws.cell(row=2, column=1).coordinate + ":" +
                          ws.cell(row=2, column=len(COLUMNS_43)).coordinate)
    for col_idx in range(1, len(COLUMNS_43) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0


# ─── SWEEP EXCEL SAVER (partial OR final) ────────────────────────────────────

def save_sweep_excel(partial_csv, output_dir, max_rows_per_sheet, mode_label,
                     is_complete, month_dir=None, ts_str=None):
    """
    Convert the partial (or complete) CSV checkpoint into a multi-sheet Excel.

    Partial output: output/full_sweep/full_sweep_partial_latest.xlsx  (fixed name,
      overwrites on every save — prevents file accumulation in git over many runs)
    Final output  : output/full_sweep/full_sweep_final_YYYYMMDD_HHMMSS.xlsx
      - Data_1 .. Data_N sheets (max_rows_per_sheet rows each, 25 000 by default)
      - Consolidated sheet: TOP 5 WinRate rows from each Data sheet

    If is_complete=True: also copies to output/YYYY-MM/Results_YYYYMMDD.xlsx
    and deletes the checkpoint CSV so the next sweep starts fresh.

    Returns the path of the saved Excel, or None if the CSV is missing.
    """
    import shutil

    if not os.path.exists(partial_csv):
        print("WARN: partial CSV not found — skipping Excel save")
        return None

    df_all = pd.read_csv(partial_csv)
    df_all = df_all.drop_duplicates(subset=["Test"], keep="last")
    df_all = df_all.sort_values("Test").reset_index(drop=True)
    total_rows = len(df_all)
    print(f"Building sweep Excel: {total_rows:,} rows, {max_rows_per_sheet:,}/sheet …")

    # Output folder: output/full_sweep/
    sweep_dir = os.path.join(output_dir, "full_sweep")
    os.makedirs(sweep_dir, exist_ok=True)

    now_ts = datetime.now()
    # FIX (space): partial uses a fixed filename (overwrites) to prevent
    # accumulation of many timestamped files in git across restart cycles.
    # Final uses a timestamp so each monthly result is archived distinctly.
    if is_complete:
        now_str  = now_ts.strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(sweep_dir, f"full_sweep_final_{now_str}.xlsx")
    else:
        out_path = os.path.join(sweep_dir, "full_sweep_partial_latest.xlsx")
    label = "final" if is_complete else "partial"

    wb = Workbook()
    sheet_num     = 1
    rows_on_sheet = 0
    sheet_top5    = {}          # sheet_name → list of row-lists for top-5 selection

    ws = wb.active
    ws.title = "Data_1"
    ws.append(COLUMNS_43)
    sheet_top5["Data_1"] = []

    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for _, r in df_all.iterrows():
        if rows_on_sheet >= max_rows_per_sheet:
            sheet_num    += 1
            ws = wb.create_sheet(title=f"Data_{sheet_num}")
            ws.append(COLUMNS_43)
            sheet_top5[f"Data_{sheet_num}"] = []
            rows_on_sheet = 0
        row_vals = [r.get(c, 0) for c in COLUMNS_43]
        ws.append(row_vals)
        sheet_top5[f"Data_{sheet_num}"].append(row_vals)
        rows_on_sheet += 1

    # Style header of every Data sheet (no per-cell zebra — too slow for 25K rows)
    now_label = now_ts.strftime("%d-%b-%Y %H:%M")
    for sn in range(1, sheet_num + 1):
        sname = f"Data_{sn}"
        if sname not in wb.sheetnames:
            continue
        ws_d = wb[sname]
        title = (f"NSE Param Sweep — {sname} — "
                 f"{total_rows:,} rows ({label}) — {now_label}")
        # Insert merged title row
        ws_d.insert_rows(1)
        ws_d.merge_cells(start_row=1, start_column=1,
                         end_row=1, end_column=len(COLUMNS_43))
        tc       = ws_d.cell(row=1, column=1)
        tc.value = title
        tc.font  = Font(bold=True, size=12, color=WHITE)
        tc.fill  = PatternFill("solid", fgColor=NAVY)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_d.row_dimensions[1].height = 24
        # Style header row (now row 2 after insert)
        for ci, col_name in enumerate(COLUMNS_43, 1):
            cell           = ws_d.cell(row=2, column=ci)
            cell.font      = Font(bold=True, size=10, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
        ws_d.row_dimensions[2].height = 18
        ws_d.freeze_panes = ws_d["A3"]
        ws_d.auto_filter.ref = (
            ws_d.cell(row=2, column=1).coordinate + ":" +
            ws_d.cell(row=2, column=len(COLUMNS_43)).coordinate
        )
        for ci in range(1, len(COLUMNS_43) + 1):
            ws_d.column_dimensions[get_column_letter(ci)].width = 13
        ws_d.page_setup.orientation = "landscape"
        ws_d.page_setup.fitToPage   = True
        ws_d.page_setup.fitToWidth  = 1
        ws_d.page_setup.fitToHeight = 0

    # ── Consolidated sheet: TOP 5 WinRate rows from each Data sheet ───────────
    winrate_idx = COLUMNS_43.index("WinRate")
    ws_cons     = wb.create_sheet(title="Consolidated")
    cons_cols   = ["SourceSheet"] + COLUMNS_43
    ws_cons.append(cons_cols)

    all_top5_count = 0
    for sname, rows in sheet_top5.items():
        if not rows:
            continue
        sorted_rows = sorted(
            rows,
            key=lambda x: float(x[winrate_idx]) if x[winrate_idx] is not None else 0,
            reverse=True
        )
        for row in sorted_rows[:5]:
            ws_cons.append([sname] + list(row))
            all_top5_count += 1

    # Style Consolidated header
    for ci, col_name in enumerate(cons_cols, 1):
        cell           = ws_cons.cell(row=1, column=ci)
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws_cons.row_dimensions[1].height = 18
    ws_cons.freeze_panes = ws_cons["A2"]
    ws_cons.auto_filter.ref = (
        ws_cons.cell(row=1, column=1).coordinate + ":" +
        ws_cons.cell(row=1, column=len(cons_cols)).coordinate
    )
    for ci in range(1, len(cons_cols) + 1):
        ws_cons.column_dimensions[get_column_letter(ci)].width = 13
    ws_cons.column_dimensions["A"].width = 15  # SourceSheet

    wb.save(out_path)
    print(f"✅ Sweep Excel saved : {out_path}")
    print(f"   Data sheets       : {sheet_num} (≤{max_rows_per_sheet:,} rows each)")
    print(f"   Consolidated      : {all_top5_count} rows (top 5 per sheet)")

    if is_complete:
        # Copy as final Results_ file in the monthly folder
        if month_dir and ts_str:
            final_path = os.path.join(month_dir, f"Results_{ts_str}.xlsx")
            shutil.copy2(out_path, final_path)
            print(f"✅ Final Results     : {final_path}")
        # Clear checkpoint so next sweep starts fresh
        try:
            os.remove(partial_csv)
            print(f"🗑  Cleared checkpoint: {partial_csv}")
        except Exception as e:
            print(f"WARN: could not remove checkpoint {partial_csv}: {e}")

    return out_path


# ─── NO-SIGNALS DIAGNOSTIC FILE ──────────────────────────────────────────────

def save_no_signals_file(cfg, mode, mode_label, symbol_filter, signals_file, month_dir, ts_str):
    prefix   = "QuickRun_Picks" if mode == "quick" else "Results"
    out_path = os.path.join(month_dir, f"{prefix}_NoSignals_{ts_str}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "No_Signals"
    ws.append(["NSE Simulation — No Signals Found"])
    ws.cell(1, 1).font = Font(bold=True, size=13, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.append([])
    ws.append(["Field", "Value"])
    for c in [ws.cell(3, 1), ws.cell(3, 2)]:
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
    q_or_p = cfg.get("quick_run" if mode == "quick" else "param_sweep", {})
    rows_d = [
        ("Run mode",          mode_label),
        ("Signal start date", cfg.get("signal_start_date", "?")),
        ("Signal end date",   cfg.get("signal_end_date",   "?")),
        ("Signals file",      signals_file),
        ("ATH min filter",    q_or_p.get("ath_min", "?")),
        ("ATH max filter",    q_or_p.get("ath_max", "?")),
        ("Pct min (5d dip)",  q_or_p.get("pct_min", "?")),
        ("Pct max (5d dip)",  q_or_p.get("pct_max", "?")),
        ("Symbol filter",     ", ".join(sorted(symbol_filter)) if symbol_filter else "ALL"),
        ("", ""),
        ("Root Cause", "ATH computed from uploaded data only. Short data -> ATH near current price."),
        ("",           "Most stocks appear near ATH and fail the -30% to -60% ATH filter."),
        ("Solution",   "Upload 1-2 years of historical NSE bhav CSVs to bhav_data/ folder."),
        ("",           "Re-run with rebuild_db=true after uploading."),
    ]
    for r in rows_d:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    wb.save(out_path)
    print(f"✅ Diagnostic file saved : {out_path}")
    return out_path


# ─── MAIN ─────────────────────────────────────────────────────────────────────

# ─── MODULE-LEVEL WORKER (must be at top level for ProcessPoolExecutor pickling) ──

def _simulate_sym_group(args):
    """
    Simulate all signals for one symbol. Defined at module level so it is
    picklable by ProcessPoolExecutor.
    """
    sym_key, sigs, sym_pd_data, params = args
    if sym_pd_data is None:
        # Symbol missing from price_dict — return invalid for each signal
        invalid = {
            "order": "Invalid", "status": None, "action": "Skip",
            "buy_count": None, "avg_buy_price": None, "total_qty": None,
            "total_investment": None, "target_price": None, "stop_price": None,
            "first_buy_date": None, "exit_found": False,
            "exit_date": None, "exit_price": None, "exit_type": None,
            "profit": None, "gain_pct": None, "market_days": None,
            "result_str": "Invalid: No market data found",
            "duration_group": None, "buy_chance": None,
            "sold_prev_close": None, "sold_open": None,
            "sold_high": None, "sold_low": None, "sold_close": None,
            "buys": [], "had_buy_chance": False,
        }
        return [dict(invalid)] * len(sigs)

    _mb, _bd, _tgt, _sl, _mdur, _inv, _fecd, _pw, _gld, _use_sl, _use_tgt = params
    pd_local = {sym_key: sym_pd_data}
    results = []
    for s in sigs:
        sim = simulate_trade_detailed(
            sym_key, s.SIGNAL_DATE, float(s.SIGNAL_CLOSE), pd_local,
            _mb, _bd, _tgt, _sl, _mdur, _inv, _fecd, _pw,
            global_last_date=_gld,
            use_stoploss=_use_sl, use_target=_use_tgt,
        )
        results.append(sim)
    return results


def main():
    parser = argparse.ArgumentParser(description="NSE Trading Simulation")
    parser.add_argument("--mode", choices=["quick", "full"], default="full",
                        help="quick = per-trade picks sheet (daily); full = aggregate stats (monthly)")
    parser.add_argument("--symbols", default="",
                        help="Comma-separated symbols e.g. TCS,WIPRO,INFY (empty = all)")
    parser.add_argument("--generate-partial-excel", action="store_true",
                        help="Only convert checkpoint CSV to partial Excel (no simulation)")
    args = parser.parse_args()
    mode = args.mode

    # ── Fast path: regenerate partial Excel from checkpoint CSV only ──────────
    # Used by the workflow's pre-commit step to ensure a fresh partial Excel
    # is always committed even when the simulation step is killed by timeout.
    if args.generate_partial_excel:
        cfg = load_config()
        max_rows_per_sheet = cfg.get("max_rows_per_sheet", 25000)
        partial_csv = os.path.join(OUTPUT_DIR, "partial", "full_sweep_partial.csv")
        if not os.path.exists(partial_csv):
            print("No checkpoint CSV found — skipping partial Excel generation.")
            return
        result = save_sweep_excel(
            partial_csv, OUTPUT_DIR, max_rows_per_sheet,
            "FULL PARAMETER SWEEP", False, None, None
        )
        if result:
            print(f"✅ Partial Excel generated: {result}")
        return

    signals_file = SIGNALS_FILE_QUICK if mode == "quick" else SIGNALS_FILE_FULL

    for f in [CONFIG_FILE, signals_file, EQ_FILE]:
        if not os.path.exists(f):
            print(f"❌ Missing: {f}")
            raise SystemExit(1)

    cfg = load_config()
    investment_per_buy       = cfg.get("investment_per_buy",       10000)
    force_exit_calendar_days = cfg.get("force_exit_calendar_days", 90)
    pending_window_days      = cfg.get("pending_window_days",       30)
    max_rows_per_sheet       = cfg.get("max_rows_per_sheet",        25000)

    trade_combos = build_trade_combos(cfg, mode)
    mode_label   = "QUICK RUN" if mode == "quick" else "FULL PARAMETER SWEEP"
    print(f"Mode            : {mode_label}")
    print(f"Trade combos    : {len(trade_combos):,}")

    # ── Symbol filter (CLI > config > all) ────────────────────────────────────
    symbol_filter = set()
    cli_syms = args.symbols.strip()
    if cli_syms:
        symbol_filter = {s.strip().upper() for s in cli_syms.split(",") if s.strip()}
        print(f"Symbol filter   : CLI override -> {sorted(symbol_filter)}")
    else:
        section     = cfg.get("quick_run" if mode == "quick" else "param_sweep", {})
        config_syms = section.get("watch_symbols", [])
        if config_syms:
            symbol_filter = {s.strip().upper() for s in config_syms if s.strip()}
            print(f"Symbol filter   : config watch_symbols -> {len(symbol_filter)} symbols")
        else:
            print("Symbol filter   : ALL EQ symbols")

    # ── Load signals ──────────────────────────────────────────────────────────
    print("Loading signals...")
    sig_df = pd.read_parquet(signals_file)
    sig_df["SIGNAL_DATE"] = pd.to_datetime(sig_df["SIGNAL_DATE"])

    if symbol_filter:
        sig_df = sig_df[sig_df["SYMBOL"].isin(symbol_filter)]
        missing = symbol_filter - set(sig_df["SYMBOL"].unique())
        if missing:
            print(f"⚠️  Symbols not in signals: {sorted(missing)}")

    print(f"  Total signals : {len(sig_df):,}")

    # ── Output path (month subfolder) ─────────────────────────────────────────
    ts_now    = datetime.now()
    ts_str    = ts_now.strftime("%Y%m%d_%H%M%S")
    month_dir = os.path.join(OUTPUT_DIR, ts_now.strftime("%Y-%m"))
    os.makedirs(month_dir, exist_ok=True)

    if len(sig_df) == 0:
        print("⚠️  No signals found. Saving diagnostic Excel.")
        save_no_signals_file(cfg, mode, mode_label, symbol_filter, signals_file, month_dir, ts_str)
        raise SystemExit(0)

    # ── Load EQ price data ────────────────────────────────────────────────────
    print("Loading EQ price data...")
    eq_df          = pd.read_parquet(EQ_FILE)
    eq_df["DATE1"] = pd.to_datetime(eq_df["DATE1"])
    if symbol_filter:
        eq_df = eq_df[eq_df["SYMBOL"].isin(symbol_filter)]
    price_dict = build_price_dict(eq_df)

    # Inject series-changed map so simulate functions can generate correct remarks
    series_changed_map = load_series_changed_map()
    if series_changed_map:
        print(f"Series-changed symbols loaded: {len(series_changed_map)} "
              f"(will be marked Invalid: Series changed from EQ to ...)")
    price_dict["__series_changed__"] = series_changed_map

    # Last data date (= "TodayDate" for non-Invalid rows)
    last_data_ts   = eq_df["DATE1"].max()
    last_data_date = last_data_ts.to_pydatetime().replace(tzinfo=None)
    print(f"Last data date  : {last_data_date.strftime('%d-%m-%Y')}")

    # For full mode: free eq_df memory immediately (not needed after price_dict)
    # For quick mode: keep eq_df alive until market data extraction below
    if mode == "full":
        del eq_df
        eq_df = None

    # ══════════════════════════════════════════════════════════════════════════
    # QUICK MODE -> Per-trade PICKS SHEET
    # ══════════════════════════════════════════════════════════════════════════
    if mode == "quick":
        q    = cfg["quick_run"]
        mb   = q["max_buys"]
        bd   = q["buy_drop"]
        tgt  = q["target"]
        sl   = q["stoploss"]
        mdur = q["max_duration"]
        use_sl  = q.get("use_stoploss", True)
        use_tgt = q.get("use_target",   True)

        print(f"Parameters      : DaysBack={q['days_back']} "
              f"PctMin={q['pct_min']:.0%} PctMax={q['pct_max']:.0%} "
              f"ATHMin={q['ath_min']:.0%} ATHMax={q['ath_max']:.0%} "
              f"MaxBuys={mb} BuyDrop={bd:.0%} Target={tgt:.0%} SL={sl:.0%} MaxDur={mdur}")
        print(f"Running picks simulation for {len(sig_df):,} signals...")

        # Deduplicate: same symbol+date -> keep first
        sig_df_picks = sig_df.drop_duplicates(
            subset=["SYMBOL", "SIGNAL_DATE"]).reset_index(drop=True)
        print(f"Unique signals  : {len(sig_df_picks):,}")

        columns  = get_picks_columns(mb)
        all_rows = []
        bought_ranges = []   # (symbol, start_ts, end_ts) for Stock History sheet
        counts   = {"Executed": 0, "Pending": 0, "Expired": 0, "Invalid": 0,
                    "Open": 0, "Closed": 0, "Profit": 0, "Loss": 0}

        # ── Build ordered signal list using itertuples (faster than iterrows) ─
        sim_params = (mb, bd, tgt, sl, mdur,
                      investment_per_buy, force_exit_calendar_days,
                      pending_window_days, last_data_date,
                      use_sl, use_tgt)

        # Group signals by symbol to pass price_dict entry once per symbol (parallel)
        sym_signal_groups = {}   # sym -> [(signal_date, signal_close, sig_tuple), ...]
        ordered_keys = []        # preserve original signal order
        for sig in sig_df_picks.itertuples(index=False):
            sym = str(sig.SYMBOL)
            if sym not in sym_signal_groups:
                sym_signal_groups[sym] = []
                ordered_keys.append(sym)
            sym_signal_groups[sym].append(sig)

        # ── Sequential simulation: one symbol group at a time ───────────────
        print(f"Simulating {len(sig_df_picks):,} signals across "
              f"{len(sym_signal_groups):,} symbols...")

        sym_sim_results = {}  # sym -> [sim, sim, ...]
        for sym in ordered_keys:
            args = (sym, sym_signal_groups[sym],
                    price_dict[sym] if sym in price_dict else None,
                    sim_params)
            sym_sim_results[sym] = _simulate_sym_group(args)

        # ── Assemble rows in original signal order ────────────────────────────
        sym_cursor = {sym: 0 for sym in ordered_keys}
        for sig in sig_df_picks.itertuples(index=False):
            sym = str(sig.SYMBOL)
            idx = sym_cursor[sym]
            sim = sym_sim_results[sym][idx]
            sym_cursor[sym] += 1

            sig_dict = sig._asdict()   # namedtuple → dict for build_picks_row
            row = build_picks_row(sig_dict, sim, price_dict, mb, last_data_date)
            all_rows.append(row)

            o = sim["order"]
            counts[o] = counts.get(o, 0) + 1
            if o == "Executed":
                s = sim["status"] or ""
                counts[s] = counts.get(s, 0) + 1
                if sim.get("result_str"):
                    rs = sim["result_str"].upper()
                    if "PROFIT" in rs: counts["Profit"] += 1
                    elif "LOSS"  in rs: counts["Loss"]  += 1

                # Track bought stocks for the BuyHistory sheet
                if sim.get("first_buy_date") is not None:
                    start_ts = pd.Timestamp(sig.SIGNAL_DATE) - pd.Timedelta(days=7)
                    end_ts   = (pd.Timestamp(sim["exit_date"])
                                if sim.get("exit_date")
                                else pd.Timestamp(last_data_date))
                    bought_ranges.append((sym, start_ts, end_ts))

        # ── Extract market data BEFORE del eq_df ──────────────────────────────
        latest_mkt_data  = build_latest_market_data(eq_df) if eq_df is not None else []
        buy_history_data = build_buy_history(eq_df, bought_ranges)
        del eq_df  # Free memory

        # ── Sort Picks by BuyDate descending (latest first, None rows at end) ─
        # BuyDate is at column index 5 (0-based) in the picks row
        # Sort: latest BuyDate first, rows with no BuyDate (Pending/Invalid) at end
        all_rows.sort(
            key=lambda r: (r[5] is not None, r[5] if r[5] is not None else datetime.min),
            reverse=True
        )

        prefix   = "QuickRun_Picks"
        out_path = os.path.join(month_dir, f"{prefix}_{ts_str}.xlsx")

        print(f"Writing {len(all_rows):,} rows to Excel (sheet: Pickse)...")
        write_picks_excel(all_rows, columns, out_path,
                          market_data=latest_mkt_data,
                          buy_history=buy_history_data)

        print(f"\n✅ Picks sheet saved : {out_path}")

        # ── Also save as fixed-path latest_quickrun.xlsx for dashboard ────────
        import shutil
        latest_path = "output/latest_quickrun.xlsx"
        shutil.copy2(out_path, latest_path)
        print(f"✅ Latest copy saved : {latest_path}")

        print(f"✅ Total signals     : {len(all_rows):,}")
        print(f"   Executed          : {counts.get('Executed',0)}  "
              f"(Open={counts.get('Open',0)}  Closed={counts.get('Closed',0)})")
        print(f"   Pending           : {counts.get('Pending',0)}")
        print(f"   Expired           : {counts.get('Expired',0)}")
        print(f"   Invalid           : {counts.get('Invalid',0)}")
        print(f"   Profit trades     : {counts.get('Profit',0)}")
        print(f"   Loss trades       : {counts.get('Loss',0)}")
        return

    # ══════════════════════════════════════════════════════════════════════════
    # FULL MODE -> Aggregate 43-column stats (parameter sweep)
    # ══════════════════════════════════════════════════════════════════════════
    filter_cols   = ["DAYSBACK", "PCTMIN", "PCTMAX", "ATHMIN", "ATHMAX"]
    filter_groups = sig_df.groupby(filter_cols, dropna=False)
    total_expected = len(filter_groups) * len(trade_combos)
    print(f"Filter combos in signals : {len(filter_groups)}")
    print(f"Total output rows        : {total_expected:,}")

    # ── Resumable checkpoint ──────────────────────────────────────────────────
    # Partial results are appended to a stable CSV after every combo so that
    # if the workflow runs out of time (GitHub Actions kills the job at the
    # timeout), the next run can pick up where the previous one left off.
    # The final formatted Excel is only written once every combo is done.
    partial_dir = os.path.join(OUTPUT_DIR, "partial")
    os.makedirs(partial_dir, exist_ok=True)
    partial_csv = os.path.join(partial_dir, "full_sweep_partial.csv")

    done_set = set()
    if os.path.exists(partial_csv):
        try:
            _done_df = pd.read_csv(partial_csv)
            done_set = set(int(t) for t in _done_df["Test"].tolist())
            print(
                f"Resume checkpoint : {len(done_set):,}/{total_expected:,} "
                f"combos already done, skipping those"
            )
            del _done_df
        except Exception as e:
            print(f"WARN: could not read partial CSV ({e}); truncating and starting fresh")
            done_set = set()
            # Re-write clean header so stale/mismatched rows don't accumulate
            pd.DataFrame(columns=COLUMNS_43).to_csv(partial_csv, index=False)

    if not os.path.exists(partial_csv):
        pd.DataFrame(columns=COLUMNS_43).to_csv(partial_csv, index=False)

    CHECKPOINT_EVERY = 5
    pending_rows     = []

    def _flush_pending():
        if not pending_rows:
            return
        pd.DataFrame(pending_rows, columns=COLUMNS_43).to_csv(
            partial_csv, mode="a", header=False, index=False
        )
        pending_rows.clear()

    test_num         = 0
    done_before      = len(done_set)
    done_this_run    = 0

    loop_start       = datetime.now()
    _last_excel_save = datetime.now()  # FIX: track last in-process Excel save time

    for filter_key, filter_group in filter_groups:
        db, pmin, pmax, amin, amax = filter_key

        # Cheap skip: if every trade combo in this filter group is already done,
        # just bump test_num and move on without materializing signal_list.
        block_start = test_num + 1
        block_end   = test_num + len(trade_combos)
        if all(t in done_set for t in range(block_start, block_end + 1)):
            test_num = block_end
            continue

        signal_list = list(zip(
            filter_group["SYMBOL"],
            filter_group["SIGNAL_DATE"],
            filter_group["SIGNAL_CLOSE"],
        ))
        total_stocks = len(signal_list)

        for (mb, bd_p, tgt, sl, mdur, use_sl, use_tgt) in trade_combos:
            test_num += 1

            if test_num in done_set:
                continue

            results = [
                simulate_trade(
                    sym, sig_date, sig_close, price_dict,
                    mb, bd_p, tgt, sl, mdur,
                    investment_per_buy, force_exit_calendar_days, pending_window_days,
                    global_last_date=last_data_date,
                    use_stoploss=use_sl, use_target=use_tgt,
                )
                for sym, sig_date, sig_close in signal_list
            ]

            stats = aggregate_stats(results)
            stats["TotalStock"] = total_stocks

            row  = [test_num, db, pmin, pmax, amin, amax, mb, bd_p, tgt, sl, mdur]
            row += [stats[c] for c in COLUMNS_43[11:]]

            pending_rows.append(row)
            done_this_run += 1

            if len(pending_rows) >= CHECKPOINT_EVERY:
                _flush_pending()
                # FIX (partial output): periodic in-process Excel save every 30 min.
                # Ensures a fresh partial Excel exists on disk even if the job is
                # later killed by the GitHub Actions timeout before main() exits.
                _now = datetime.now()
                if (_now - _last_excel_save).total_seconds() >= 1800:
                    print("  [Periodic save] Generating partial Excel from checkpoint…")
                    save_sweep_excel(
                        partial_csv, OUTPUT_DIR, max_rows_per_sheet,
                        mode_label, False, None, None
                    )
                    _last_excel_save = _now

        if (done_before + done_this_run) % 200 == 0 and done_this_run:
            elapsed = (datetime.now() - loop_start).total_seconds()
            rate    = done_this_run / max(elapsed, 1)
            remaining = total_expected - (done_before + done_this_run)
            eta_min   = remaining / rate / 60 if rate else 0
            print(
                f"  Completed {done_before + done_this_run:,}/{total_expected:,} "
                f"(this run: {done_this_run:,}, rate: {rate:.2f}/s, "
                f"ETA: {eta_min:.1f} min)"
            )

    _flush_pending()

    total_done  = done_before + done_this_run
    is_complete = (total_done >= total_expected)

    # Always save sweep Excel from checkpoint (partial OR complete)
    sweep_excel = save_sweep_excel(
        partial_csv, OUTPUT_DIR, max_rows_per_sheet, mode_label,
        is_complete, month_dir, ts_str
    )

    if not is_complete:
        print(
            f"\n⏸  Partial run: {total_done:,}/{total_expected:,} "
            f"combos complete ({done_this_run:,} this run)."
        )
        print(f"   Checkpoint : {partial_csv}")
        if sweep_excel:
            print(f"   Excel saved: {sweep_excel}")
        print(f"   Re-run the 'Monthly Parameter Sweep' workflow to resume from here.")
        return

    print(f"\n✅ All {total_expected:,} combos complete!")
    if sweep_excel:
        print(f"✅ Final Excel    : {sweep_excel}")


if __name__ == "__main__":
    start = datetime.now()
    main()
    elapsed = (datetime.now() - start).total_seconds()
    print(f"\n⏱  Total time: {elapsed:.1f}s ({elapsed/60:.1f} min)")
