"""
Script 6: Consolidate Result Excel Files
==========================================
Reads:  All *.xlsx files in output/ folder (skips ~temp files, skips Consolidated_*)
Merges: All data into one professional Excel file
Output: output/Consolidated_YYYYMMDD_HHMMSS.xlsx

43-column fixed order (case-insensitive match):
  Test, DAYSBACK, PCTMIN, PCTMAX, ATHMIN, ATHMAX, MAXBUYS, BUYDROP,
  TARGET, STOPLOSS, MAXDURA, WinRate, TotalTrade, Executed, Open, Closed,
  ProfitTGT, LossSL, LossFEMD, LossFECD, ProfitFEMD, ProfitFECD,
  Pending, Expired, Invalid, TotalRows, Wins, Losses, TotalStock,
  SumProfit, SumGainFin, Dur5, Dur10, Dur15, Dur20, Dur25, Dur30,
  Dur35, Dur40, ExitTGT, ExitSL, ExitFEMD, ExitFECD

Features:
  - Header from first file only; drop duplicate rows
  - 500,000 rows per sheet max; overflow → Data_2, Data_3...
  - Professional formatting: merged title, navy headers, zebra stripes,
    borders, freeze panes, autofilter, landscape print
"""

import os
import glob
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR      = "output"
MAX_ROWS_SHEET  = 500_000
NAVY            = "00203864"
WHITE           = "00FFFFFF"
LIGHT           = "00F2F2F2"

COLUMNS_43 = [
    "Test", "DAYSBACK", "PCTMIN", "PCTMAX", "ATHMIN", "ATHMAX",
    "MAXBUYS", "BUYDROP", "TARGET", "STOPLOSS", "MAXDURA",
    "WinRate", "TotalTrade", "Executed", "Open", "Closed",
    "ProfitTGT", "LossSL", "LossFEMD", "LossFECD", "ProfitFEMD", "ProfitFECD",
    "Pending", "Expired", "Invalid", "TotalRows", "Wins", "Losses", "TotalStock",
    "SumProfit", "SumGainFin",
    "Dur5", "Dur10", "Dur15", "Dur20", "Dur25", "Dur30", "Dur35", "Dur40",
    "ExitTGT", "ExitSL", "ExitFEMD", "ExitFECD"
]
COLUMNS_UPPER = [c.upper() for c in COLUMNS_43]

def read_excel_sheets(filepath):
    """Read all data sheets from a result Excel file."""
    frames = []
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            # Find header row (first non-empty row)
            header_row = None
            data_start = 0
            for idx, row in enumerate(rows):
                if row and any(v is not None for v in row):
                    # Check if it looks like a header (contains column names)
                    row_upper = [str(v).upper().strip() if v else "" for v in row]
                    if "TEST" in row_upper or "WINRATE" in row_upper:
                        header_row = row_upper
                        data_start = idx + 1
                        break
            if header_row is None:
                continue

            data_rows = rows[data_start:]
            if not data_rows:
                continue

            df = pd.DataFrame(data_rows, columns=header_row)
            frames.append(df)
        wb.close()
    except Exception as e:
        print(f"  ⚠️  Error reading {filepath}: {e}")
    return frames

def reorder_columns(df):
    """Reorder to 43-column fixed order; missing cols → 0; extra cols ignored."""
    result = {}
    df_upper = {c.upper(): c for c in df.columns}
    for col in COLUMNS_43:
        src = df_upper.get(col.upper())
        if src:
            result[col] = df[src].values
        else:
            result[col] = [0] * len(df)
    return pd.DataFrame(result)

def apply_formatting(ws, total_rows):
    """Apply professional formatting to a worksheet."""
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = len(COLUMNS_43)

    # Insert title row at top
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value     = f"NSE Simulation — Consolidated Results | {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    title_cell.font      = Font(bold=True, size=14, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row (now row 2)
    for col_idx, col_name in enumerate(COLUMNS_43, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = col_name
        cell.font      = Font(bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
    ws.row_dimensions[2].height = 18

    # Zebra stripe data rows
    last_row = total_rows + 2      # +2 for title + header
    for row in range(3, last_row + 1):
        fill_color = LIGHT if row % 2 == 0 else "00FFFFFF"
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor.rgb != NAVY:   # don't overwrite header
                cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")

    # Freeze panes
    ws.freeze_panes = ws.cell(row=3, column=1)

    # Autofilter
    ws.auto_filter.ref = (ws.cell(row=2, column=1).coordinate + ":" +
                          ws.cell(row=2, column=n_cols).coordinate)

    # Column widths
    for col_idx in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

def main():
    # Find source files: *.xlsx in output/, skip ~temp, skip Consolidated_*
    pattern = os.path.join(OUTPUT_DIR, "*.xlsx")
    all_files = glob.glob(pattern)
    source_files = [
        f for f in sorted(all_files)
        if not os.path.basename(f).startswith("~")
        and "consolidated" not in os.path.basename(f).lower()
    ]

    print(f"Source files found: {len(source_files)}")
    if not source_files:
        print("❌ No result files found in output/")
        raise SystemExit(1)

    all_frames = []
    for fp in source_files:
        print(f"  Reading: {os.path.basename(fp)}")
        sheets = read_excel_sheets(fp)
        all_frames.extend(sheets)

    if not all_frames:
        print("❌ No data extracted from source files.")
        raise SystemExit(1)

    print("Merging data...")
    combined = pd.concat(all_frames, ignore_index=True)
    before = len(combined)
    combined.drop_duplicates(inplace=True)
    after = len(combined)
    print(f"Rows: {before:,} → {after:,} (removed {before - after:,} duplicates)")

    # Reorder to 43-column fixed order
    combined = reorder_columns(combined)

    # Re-number Test column
    combined["Test"] = range(1, len(combined) + 1)

    ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUTPUT_DIR, f"Consolidated_{ts_str}.xlsx")

    print("Writing Excel output...")
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Data_1"
    ws.append(COLUMNS_43)

    sheet_num     = 1
    rows_on_sheet = 0
    sheet_row_counts = {1: 0}

    for _, row_data in combined.iterrows():
        if rows_on_sheet >= MAX_ROWS_SHEET:
            sheet_num += 1
            ws = wb.create_sheet(title=f"Data_{sheet_num}")
            ws.append(COLUMNS_43)
            rows_on_sheet = 0
            sheet_row_counts[sheet_num] = 0

        ws.append([row_data[c] for c in COLUMNS_43])
        rows_on_sheet += 1
        sheet_row_counts[sheet_num] = rows_on_sheet

    print(f"Applying formatting to {sheet_num} sheet(s)...")
    for sn in range(1, sheet_num + 1):
        sname = f"Data_{sn}"
        if sname in wb.sheetnames:
            apply_formatting(wb[sname], sheet_row_counts.get(sn, 0))

    wb.save(out_path)
    print(f"\n✅ Consolidated file saved : {out_path}")
    print(f"✅ Total rows              : {len(combined):,}")
    print(f"✅ Sheets                  : {sheet_num}")

if __name__ == "__main__":
    main()
